from PyQt5 import QtCore, QtGui, QtWidgets, Qt
from PyQt5.QtWidgets import QMessageBox
import sqlite3
import hashlib
from PyQt5.QtCore import *
import openpyxl
from openpyxl import load_workbook
from table import *
import os
## pyinstaller --onefile --noconsole main.py



class Ui_MainWindow(object):


    # Возвращаем список
    def sql_return_table_in_list(self, sql_command):
        connection = sqlite3.connect("Store.db")
        cursor = connection.cursor()
        sql = sql_command
        cursor.execute(sql)
        result = cursor.fetchall()
        connection.commit()
        connection.close()
        return result

    def sql_just_execution_of_a_command(self, sql_command):
        connection = sqlite3.connect("Store.db")
        cursor = connection.cursor()
        sql = sql_command
        cursor.execute(sql)
        connection.commit()
        connection.close()

    # Переключение между страницами
    def set_current_index1(self):
        self.tabWidget.setCurrentIndex(0)
        self.label_tab1_login_error_massage.setText(" ")

    def set_current_index2(self):
        self.tabWidget.setCurrentIndex(1)
        self.label_tab2_error.setText(' ')

    def set_current_index3(self):
        self.tabWidget.setCurrentIndex(2)

    def set_current_index4(self):
        self.tabWidget.setCurrentIndex(3)

    def set_current_index5(self):
        self.tabWidget.setCurrentIndex(4)

    def set_current_index6(self):
        self.tabWidget.setCurrentIndex(5)

    def set_current_index7(self):
        self.tabWidget.setCurrentIndex(6)

    # вывод всплывающего окна.
    def show_message(self, title, message):
        msg = QMessageBox()
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setStyleSheet("background-color: #2D364C;\n"
                                 "color: rgb(255, 255, 255);")
        x = msg.exec_()

    # Регистрация
    def registation_new_user(self):
        self.label_tab1_login_error_massage.clear()
        self.label_tab2_error.clear()
        login = self.lineEdit_part2_input_login.text().capitalize()
        password = self.lineEdit_part2_input_password.text().capitalize()
        double_password = self.lineEdit_part2_input_password_double.text().capitalize()
        nick = self.lineEdit_part2_input_nameAndSurName.text().capitalize()

        if password != double_password:
            self.label_tab2_error.setText('Пароль не совпадает')
            self.lineEdit_part2_input_password.clear()
            self.lineEdit_part2_input_password_double.clear()
            return 0

        if len(login) < 4 or len(password) < 4 or len(nick) < 4:
            self.label_tab2_error.setText('Минимальная длинна 4 символа')
            self.lineEdit_part2_input_password.clear()
            self.lineEdit_part2_input_password_double.clear()
            self.lineEdit_part2_input_login.clear()
            self.lineEdit_part2_input_nameAndSurName.clear()
            return 0

        login = hashlib.sha3_256(login.encode()).hexdigest()
        password = hashlib.sha3_256(password.encode()).hexdigest()

        sql = f"SELECT * FROM users WHERE login_user_in_users = '{login}'"
        result = self.sql_return_table_in_list(sql)
        if len(result) > 0:
            self.label_tab2_error.setText('Пользователь уже существует')
            self.label_tab2_error.setText('Логин занят')
            return 0
        sql = f"""INSERT INTO users (login_user_in_users, password_user_in_users, name_surname_in_users)
            VALUES('{login}', '{password}', '{nick}');"""
        self.sql_just_execution_of_a_command(sql)
        self.label_tab2_error.setText('Аккаунт успешно создан')


    # авторизация
    def join_in_programm(self):
        self.label_tab1_login_error_massage.clear()
        login = self.lineEdit_part1_login.text().capitalize()
        password = self.lineEdit_part2_password.text().capitalize()


        login_hash = hashlib.sha3_256(login.encode()).hexdigest()
        password_hash = hashlib.sha3_256(password.encode()).hexdigest()
        sql = f"SELECT id_user_in_users,name_surname_in_users FROM users WHERE login_user_in_users = '{login_hash}' AND password_user_in_users = '{password_hash}';"
        result = self.sql_return_table_in_list(sql)

        if len(result) == 0:
            self.lineEdit_part1_login.clear()
            self.lineEdit_part2_password.clear()
            self.label_tab1_login_error_massage.setText('Не верные данные')
            return 0
        else:
            self.pushButton_de_login.setVisible(True)
            self.pushButton_main_work_spase_go_part3.setEnabled(True)
            self.pushButton_report_go_part4.setEnabled(True)
            self.pushButton_rating_workers_go_part5.setEnabled(True)
            self.pushButton_report_for_print_go_part6.setEnabled(True)
            login = self.lineEdit_part1_login.text()
            login = result[0][1]
            self.label_you_are_welcom_name.setText(login)
            self.viev_table()
            self.tabWidget.setCurrentIndex(2)



    # де авторизация
    def delogin_in_programm(self):
        self.tabWidget.setCurrentIndex(0)
        self.pushButton_de_login.setVisible(False)
        self.pushButton_main_work_spase_go_part3.setEnabled(False)
        self.pushButton_report_go_part4.setEnabled(False)
        self.pushButton_rating_workers_go_part5.setEnabled(False)
        self.pushButton_report_for_print_go_part6.setEnabled(False)
        self.lineEdit_part1_login.clear()
        self.lineEdit_part2_password.clear()
        self.label_you_are_welcom_name.clear()
        # забыть логин,


    ### 4 Добавление записи+
    def add_in_work_spase(self):
        name_product = self.comboBox_part3_product_name.currentText()
        name_organization = self.comboBox_tab3_name_orhanization.currentText()
        count_product = self.spinBox_part3_count_product.value()
        if self.radioButton_part3_minus.isChecked():
            count_product = -count_product
        if name_product == 'Имя продукта' or name_organization == 'Имя организации':
            self.show_message('Ошибка', 'Имя продукта и имя организации обязательное поле')
            return 0

        sql = f"""SELECT id_product from product where product.name_product = '{name_product}'"""
        result = self.sql_return_table_in_list(sql)
        id_product = result[0][0]

        sql = f"SELECT id_organization from organization where organization.name_organization = '{name_organization}'"
        result = self.sql_return_table_in_list(sql)
        id_organization = result[0][0]

        nick = self.label_you_are_welcom_name.text()
        nick = hashlib.sha3_256(nick.encode()).hexdigest()
        sql = f"select id_user_in_users from users where login_user_in_users = '{nick}'"
        result = self.sql_return_table_in_list(sql)
        nick = str(result[0][0])
        sql = f'''INSERT INTO movements (id_product, count, id_organization, data, id_who_made_movement) 
        VALUES ({id_product}, {count_product}, {id_organization}, date('now'), '{nick}')'''
        self.sql_just_execution_of_a_command(sql)
        self.viev_table()


    # удаление записей
    def delete_from_sql_movements(self):
        current_row = self.tableWidget_part3_work_spase.currentRow()
        if current_row < 0:
            self.show_message("оповещение", "Для удаления выберите строчку.")
            return 0
        id_movement = self.tableWidget_part3_work_spase.item(current_row, 0)
        id_movement = int(id_movement.data(Qt.DisplayRole))
        self.show_message('оповещение', f'Запись с номером перемещения {id_movement} удалена!. Все удаленные записи сохраняются в лог!')
        sql = f"DELETE FROM movements WHERE id_movement = {id_movement}"
        self.sql_just_execution_of_a_command(sql)
        self.viev_table()


    # изменение записи
    def update_from_sql_movements(self):
        current_row = self.tableWidget_part3_work_spase.currentRow()
        if current_row < 0:
            self.show_message("уведомление",
                              "для изменения выберете нужные значения а затем кликните на нужную строчку")
            return 0
        id_movement = self.tableWidget_part3_work_spase.item(current_row, 0)
        id_movement = int(id_movement.data(Qt.DisplayRole))

        name_product = self.comboBox_part3_product_name.currentText()
        name_organization = self.comboBox_tab3_name_orhanization.currentText()
        count_product = self.spinBox_part3_count_product.value()
        if self.radioButton_part3_minus.isChecked():
            count_product = -count_product

        if name_product == 'Имя продукта' or name_organization == 'Имя организации':
            self.show_message('Ошибка', 'Имя продукта и имя организации обязательное поле')
            return 0

        sql = f"""SELECT id_product from product where product.name_product = '{name_product}'"""
        result = self.sql_return_table_in_list(sql)
        id_product = result[0][0]
        sql = f"SELECT id_organization from organization where organization.name_organization = '{name_organization}'"
        result = self.sql_return_table_in_list(sql)
        id_organization = result[0][0]
        sql = f'''UPDATE movements 
            SET id_product = '{id_product}', 
                count = "{count_product}", 
                id_organization = "{id_organization}"
                WHERE id_movement = {id_movement}'''
        self.sql_just_execution_of_a_command(sql)
        self.viev_table()
        self.show_message("уведомление", f"Запись с перемещением №{id_movement} изменена, все данные сохранены в историю изменений!")

    ### Заполнение таблицы
    def viev_table(self):
        connection = sqlite3.connect("Store.db")
        cursor = connection.cursor()
        sql = """SELECT movements.id_movement, product.name_product , movements.count, organization.name_organization, movements.data, users.name_surname_in_users
        FROM movements 
        INNER JOIN product on product.id_product = movements.id_product 
        INNER JOIN organization on organization.id_organization = movements.id_organization 
        INNER JOIN users on users.id_user_in_users = movements.id_who_made_movement"""

        result = cursor.execute(sql)
        self.tableWidget_part3_work_spase.setRowCount(0)
        for row_number, row_data in enumerate(result):
            self.tableWidget_part3_work_spase.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.tableWidget_part3_work_spase.setItem(row_number, column_number,
                                                          QtWidgets.QTableWidgetItem(str(data)))
        connection.commit()
        connection.close()  # add
        self.tableWidget_part3_work_spase.resizeColumnsToContents()

    ######################################################################## Страница 4

    # Вывод отчета по конкретной организации
    def viev_otchet_organization(self):
        name_organization = self.comboBox_tab4_name_organization.currentText()

        if name_organization == 'Имя организации':
            self.show_message('уведомление', 'Нужно выбрать организацию')
            return 0

        sql = f"""SELECT movements.id_movement, product.name_product, movements.count,organization.name_organization,movements.data,users.name_surname_in_users
        FROM organization,movements, product, users 
        WHERE organization.name_organization = '{name_organization}'
        AND movements.id_organization = organization.id_organization
        AND movements.id_product = product.id_product
        AND movements.id_who_made_movement = id_user_in_users"""
        result = self.sql_return_table_in_list(sql)
        if len(result) < 1:
            result.append(['Нету', 'данных'])
        self.tableWidget_tab4.setRowCount(0)
        for row_number, row_data in enumerate(result):
            self.tableWidget_tab4.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.tableWidget_tab4.setItem(row_number, column_number,
                                              QtWidgets.QTableWidgetItem(str(data)))
        self.tableWidget_tab4.resizeColumnsToContents()


    ##вывод отчета по конкретному имени.
    def viev_otchet_product(self):
        name_product = self.comboBox_tab5_name_product.currentText()
        if name_product == "Все товары":
            sql = """SELECT product.name_product, sum(movements.count), count(movements.count), max(data)
            FROM product, movements
            where movements.id_product = product.id_product
            GROUP by product.name_product"""
        else:
            sql = f"""SELECT product.name_product, sum(movements.count), count(movements.count), max(data)
                    FROM movements,product 
                    WHERE name_product = '{name_product}' 
                    AND movements.id_product = product.id_product"""
        result = self.sql_return_table_in_list(sql)
        self.tableWidget_tab5.setRowCount(0)
        for row_number, row_data in enumerate(result):
            self.tableWidget_tab5.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.tableWidget_tab5.setItem(row_number, column_number,
                                              QtWidgets.QTableWidgetItem(str(data)))
        self.tableWidget_tab5.resizeColumnsToContents()


    def report_from_exel(self):
        sql = """SELECT movements.id_movement, product.name_product , movements.count, organization.name_organization, movements.data, users.name_surname_in_users
        FROM movements 
        INNER JOIN product on product.id_product = movements.id_product 
        INNER JOIN organization on organization.id_organization = movements.id_organization 
        INNER JOIN users on users.id_user_in_users = movements.id_who_made_movement"""
        result = self.sql_return_table_in_list(sql)
        if len(result) < 1:
            self.show_message('ошибка', 'Нет записей в базе данных')
            return 0
        filepath = "./report.xlsx"
        try:
            os.remove(filepath)  # удаляем старый файл
        except FileNotFoundError:
            pass

        wb = openpyxl.Workbook()
        wb.save(filepath)
        wb = load_workbook(filepath)
        ws = wb['Sheet']
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25
        result.insert(0, ['id', 'имя товара', 'кол-во', 'организация', 'дата', 'сотрудник'])
        for row in range(1, len(result)+1):
            for col in range(1, len(result[1])+1):
                value = result[row-1][col-1]
                cell = ws.cell(row=row, column=col)
                cell.value = value
        wb.save(filepath)
        wb.close()
        self.show_message("Уведомление", "отчет сформирован и сохранен в папке с программой")



    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.setFixedSize(1169, 795)
        MainWindow.setStyleSheet("background-color: #2D364C;\n"
                                 "color: rgb(255, 255, 255);")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setGeometry(QtCore.QRect(200, 100, 961, 641))
        self.tabWidget.setStyleSheet("QPushButton{\n"
                                     "border-radius: 13;\n"
                                     "border: 2px solid rgb(255, 255, 255);\n"
                                     "}\n"
                                     "QLineEdit{\n"
                                     "border-radius: 13;\n"
                                     "border: 2px solid rgb(255, 255, 255);\n"
                                     "}\n"
                                     "QLineEdit:hover{\n"
                                     "border-radius: 13;\n"
                                     "border: 2px solid rgb(113, 215, 78);\n"
                                     "}\n"
                                     "QPushButton:hover{\n"
                                     "background-color: rgb(113, 215, 78);\n"
                                     "border-radius: 13;\n"
                                     "}\n"
                                     "")
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.label_3_dont_touch = QtWidgets.QLabel(self.tab)
        self.label_3_dont_touch.setGeometry(QtCore.QRect(280, 150, 421, 51))
        font = QtGui.QFont()
        font.setPointSize(27)
        self.label_3_dont_touch.setFont(font)
        self.label_3_dont_touch.setObjectName("label_3_dont_touch")
        self.label_4_dont_touch = QtWidgets.QLabel(self.tab)
        self.label_4_dont_touch.setGeometry(QtCore.QRect(490, 200, 191, 16))
        font = QtGui.QFont()
        font.setPointSize(7)
        font.setBold(False)
        font.setWeight(50)
        self.label_4_dont_touch.setFont(font)
        self.label_4_dont_touch.setStyleSheet("color: rgb(113, 215, 78)")
        self.label_4_dont_touch.setObjectName("label_4_dont_touch")
        self.lineEdit_part1_login = QtWidgets.QLineEdit(self.tab)
        self.lineEdit_part1_login.setGeometry(QtCore.QRect(300, 240, 281, 51))
        font = QtGui.QFont()
        font.setPointSize(10)

        rx = QtCore.QRegExp("[a-zA-Z0-9.,]{100}")
        validator = QtGui.QRegExpValidator(rx) #добавил валидатор

        self.lineEdit_part1_login.setFont(font)
        self.lineEdit_part1_login.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_part1_login.setObjectName("lineEdit_part1_login")
        self.lineEdit_part1_login.setValidator(validator)

        self.lineEdit_part2_password = QtWidgets.QLineEdit(self.tab)
        self.lineEdit_part2_password.setGeometry(QtCore.QRect(300, 310, 281, 51))
        self.lineEdit_part2_password.setEchoMode(QtWidgets.QLineEdit.Password)
        self.lineEdit_part2_password.setValidator(validator)

        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(10)
        self.lineEdit_part2_password.setFont(font)
        self.lineEdit_part2_password.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_part2_password.setObjectName("lineEdit_part2_password")
        self.pushButton_part2_autorization_join = QtWidgets.QPushButton(self.tab)
        self.pushButton_part2_autorization_join.setGeometry(QtCore.QRect(350, 410, 181, 51))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.pushButton_part2_autorization_join.setFont(font)
        self.pushButton_part2_autorization_join.setStyleSheet("border-radius: 13;\n"
                                                              "border: 2px solid rgb(255, 255, 255);")
        self.pushButton_part2_autorization_join.setObjectName("pushButton_part2_autorization_join")
        self.pushButton_part2_registration_go_to_part2 = QtWidgets.QPushButton(self.tab)
        self.pushButton_part2_registration_go_to_part2.setGeometry(QtCore.QRect(560, 130, 111, 31))
        font = QtGui.QFont()
        font.setPointSize(7)
        self.pushButton_part2_registration_go_to_part2.setFont(font)
        self.pushButton_part2_registration_go_to_part2.setObjectName("pushButton_part2_registration_go_to_part2")
        self.label_tab1_login_error_massage = QtWidgets.QLabel(self.tab)
        self.label_tab1_login_error_massage.setGeometry(QtCore.QRect(300, 380, 321, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_tab1_login_error_massage.setFont(font)
        self.label_tab1_login_error_massage.setStyleSheet("color: rgb(255, 0, 4);")
        self.label_tab1_login_error_massage.setText("")
        self.label_tab1_login_error_massage.setObjectName("label_tab1_login_error_massage")
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.lineEdit_part2_input_password = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_part2_input_password.setGeometry(QtCore.QRect(290, 190, 301, 51))
        self.lineEdit_part2_input_password.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_part2_input_password.setObjectName("lineEdit_part2_input_password")
        self.lineEdit_part2_input_password.setEchoMode(QtWidgets.QLineEdit.Password)
        self.lineEdit_part2_input_password_double = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_part2_input_password_double.setGeometry(QtCore.QRect(290, 270, 301, 51))
        self.lineEdit_part2_input_password_double.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_part2_input_password_double.setObjectName("lineEdit_part2_input_password_double")
        self.lineEdit_part2_input_password_double.setEchoMode(QtWidgets.QLineEdit.Password)
        self.lineEdit_part2_input_password_double.setValidator(validator)
        self.lineEdit_part2_input_password.setValidator(validator)

        self.lineEdit_part2_input_nameAndSurName = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_part2_input_nameAndSurName.setGeometry(QtCore.QRect(290, 350, 301, 51))
        self.lineEdit_part2_input_nameAndSurName.setText("")
        self.lineEdit_part2_input_nameAndSurName.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_part2_input_nameAndSurName.setObjectName("lineEdit_part2_input_nameAndSurName")
        self.lineEdit_part2_input_nameAndSurName.setValidator(validator)
        self.pushButton_part2_registration = QtWidgets.QPushButton(self.tab_2)
        self.pushButton_part2_registration.setGeometry(QtCore.QRect(290, 450, 171, 61))
        self.pushButton_part2_registration.setObjectName("pushButton_part2_registration")
        self.pushButton_part2_go_to_part1 = QtWidgets.QPushButton(self.tab_2)
        self.pushButton_part2_go_to_part1.setGeometry(QtCore.QRect(480, 450, 111, 61))
        self.pushButton_part2_go_to_part1.setObjectName("pushButton_part2_go_to_part1")
        self.lineEdit_part2_input_login = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_part2_input_login.setGeometry(QtCore.QRect(290, 120, 301, 51))
        self.lineEdit_part2_input_login.setAlignment(QtCore.Qt.AlignCenter)
        self.lineEdit_part2_input_login.setObjectName("lineEdit_part2_input_login")
        self.lineEdit_part2_input_login.setValidator(validator)

        self.label_17_dont_touch = QtWidgets.QLabel(self.tab_2)
        self.label_17_dont_touch.setGeometry(QtCore.QRect(340, 40, 201, 61))
        font = QtGui.QFont()
        font.setPointSize(20)
        self.label_17_dont_touch.setFont(font)
        self.label_17_dont_touch.setStyleSheet("color: rgb(113, 215, 78);")
        self.label_17_dont_touch.setObjectName("label_17_dont_touch")
        self.label_tab2_error = QtWidgets.QLabel(self.tab_2)
        self.label_tab2_error.setGeometry(QtCore.QRect(296, 420, 291, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_tab2_error.setFont(font)
        self.label_tab2_error.setStyleSheet("color: rgb(255, 0, 4);")
        self.label_tab2_error.setText("")
        self.label_tab2_error.setObjectName("label_tab2_error")
        self.tabWidget.addTab(self.tab_2, "")
        self.tab_3_ = QtWidgets.QWidget()
        self.tab_3_.setObjectName("tab_3_")
        self.spinBox_part3_count_product = QtWidgets.QSpinBox(self.tab_3_)
        self.spinBox_part3_count_product.setGeometry(QtCore.QRect(410, 30, 81, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.spinBox_part3_count_product.setFont(font)
        self.spinBox_part3_count_product.setObjectName("spinBox_part3_count_product")
        self.spinBox_part3_count_product.setMinimum(1)
        self.comboBox_part3_product_name = QtWidgets.QComboBox(self.tab_3_)
        self.comboBox_part3_product_name.setGeometry(QtCore.QRect(120, 30, 281, 31))
        self.comboBox_part3_product_name.setObjectName("comboBox_part3_product_name")
        self.comboBox_part3_product_name.addItem("Имя продукта")
        self.pushButton_part3_add_in_work_spase = QtWidgets.QPushButton(self.tab_3_)
        self.pushButton_part3_add_in_work_spase.setGeometry(QtCore.QRect(270, 80, 361, 41))
        self.pushButton_part3_add_in_work_spase.setObjectName("pushButton_part3_add_in_work_spase")
        self.radioButton_part3_plus = QtWidgets.QRadioButton(self.tab_3_)
        self.radioButton_part3_plus.setChecked(True)
        self.radioButton_part3_plus.setGeometry(QtCore.QRect(20, 20, 95, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.radioButton_part3_plus.setFont(font)
        self.radioButton_part3_plus.setObjectName("radioButton_part3_plus")
        self.radioButton_part3_minus = QtWidgets.QRadioButton(self.tab_3_)
        self.radioButton_part3_minus.setGeometry(QtCore.QRect(20, 40, 95, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.radioButton_part3_minus.setFont(font)
        self.radioButton_part3_minus.setObjectName("radioButton_part3_minus")

        self.tableWidget_part3_work_spase = QtWidgets.QTableWidget(self.tab_3_)
        self.tableWidget_part3_work_spase.setGeometry(QtCore.QRect(0, 150, 951, 461))
        self.tableWidget_part3_work_spase.setObjectName("tableWidget_part3_work_spase")
        self.tableWidget_part3_work_spase.setColumnCount(6)
        self.tableWidget_part3_work_spase.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.tableWidget_part3_work_spase.setStyleSheet("color: rgb(113, 215, 78);")  ## Поменял цвет плохо видно было
        self.label_18_dont_touch = QtWidgets.QLabel(self.tab_3_)
        self.label_18_dont_touch.setGeometry(QtCore.QRect(240, 10, 55, 16))
        self.label_18_dont_touch.setObjectName("label_18_dont_touch")
        self.label_19_dont_touch = QtWidgets.QLabel(self.tab_3_)
        self.label_19_dont_touch.setGeometry(QtCore.QRect(420, 10, 55, 16))
        self.label_19_dont_touch.setObjectName("label_19_dont_touch")
        self.pushButton_part3_delete = QtWidgets.QPushButton(self.tab_3_)
        self.pushButton_part3_delete.setGeometry(QtCore.QRect(840, 40, 101, 28))
        self.pushButton_part3_delete.setObjectName("pushButton_part3_delete")
        self.pushButton_part3_edit = QtWidgets.QPushButton(self.tab_3_)
        self.pushButton_part3_edit.setGeometry(QtCore.QRect(840, 10, 101, 28))
        self.pushButton_part3_edit.setObjectName("pushButton_part3_edit")
        self.comboBox_tab3_name_orhanization = QtWidgets.QComboBox(self.tab_3_)
        self.comboBox_tab3_name_orhanization.setGeometry(QtCore.QRect(520, 30, 301, 31))
        self.comboBox_tab3_name_orhanization.setObjectName("comboBox_tab3_name_orhanization")
        self.comboBox_tab3_name_orhanization.addItem("")
        self.label_18_dont_touch_ = QtWidgets.QLabel(self.tab_3_)
        self.label_18_dont_touch_.setGeometry(QtCore.QRect(630, 10, 71, 16))
        self.label_18_dont_touch_.setObjectName("label_18_dont_touch_")
        self.label_tab3_error = QtWidgets.QLabel(self.tab_3_)
        self.label_tab3_error.setGeometry(QtCore.QRect(640, 80, 271, 31))
        self.label_tab3_error.setStyleSheet("color: rgb(255, 0, 4);")
        self.label_tab3_error.setText("")
        self.label_tab3_error.setObjectName("label_tab3_error")
        self.tabWidget.addTab(self.tab_3_, "")
        self.tab_4_ = QtWidgets.QWidget()
        self.tab_4_.setObjectName("tab_4_")
        self.tableWidget_tab4 = QtWidgets.QTableWidget(self.tab_4_)
        self.tableWidget_tab4.setGeometry(QtCore.QRect(0, 100, 951, 511))
        self.tableWidget_tab4.setObjectName("tableWidget_tab4")
        self.tableWidget_tab4.setStyleSheet("color: rgb(113, 215, 78);")  ## Поменял цвет плохо видно было
        self.tableWidget_tab4.setColumnCount(6)
        self.tableWidget_tab4.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)

        self.pushButton_tab4_view_report = QtWidgets.QPushButton(self.tab_4_)
        self.pushButton_tab4_view_report.setGeometry(QtCore.QRect(570, 30, 171, 51))
        self.pushButton_tab4_view_report.setObjectName("pushButton_tab4_view_report")
        self.comboBox_tab4_name_organization = QtWidgets.QComboBox(self.tab_4_)
        self.comboBox_tab4_name_organization.setGeometry(QtCore.QRect(120, 40, 411, 31))
        self.comboBox_tab4_name_organization.setObjectName("comboBox_tab4_name_organization")
        self.comboBox_tab4_name_organization.addItem("")
        self.label_dont_touch_2 = QtWidgets.QLabel(self.tab_4_)
        self.label_dont_touch_2.setGeometry(QtCore.QRect(160, 20, 361, 16))
        self.label_dont_touch_2.setObjectName("label_dont_touch_2")
        self.tabWidget.addTab(self.tab_4_, "")
        self.tab_5_ = QtWidgets.QWidget()
        self.tab_5_.setObjectName("tab_5_")

        self.tableWidget_tab5 = QtWidgets.QTableWidget(self.tab_5_)
        self.tableWidget_tab5.setGeometry(QtCore.QRect(0, 100, 951, 511))
        self.tableWidget_tab5.setObjectName("tableWidget_tab5")
        self.tableWidget_tab5.setStyleSheet("color: rgb(113, 215, 78);")  ## Поменял цвет плохо видно было
        self.tableWidget_tab5.setColumnCount(4)
        self.tableWidget_tab5.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)

        self.comboBox_tab5_name_product = QtWidgets.QComboBox(self.tab_5_)
        self.comboBox_tab5_name_product.setGeometry(QtCore.QRect(150, 50, 341, 22))
        self.comboBox_tab5_name_product.setObjectName("comboBox_tab5_name_product")
        self.comboBox_tab5_name_product.addItem("")
        self.pushButton_tab5_view_report = QtWidgets.QPushButton(self.tab_5_)
        self.pushButton_tab5_view_report.setGeometry(QtCore.QRect(560, 50, 131, 23))
        self.pushButton_tab5_view_report.setObjectName("pushButton_tab5_view_report")
        self.tabWidget.addTab(self.tab_5_, "")
        self.tab_6_report = QtWidgets.QWidget()
        self.tab_6_report.setObjectName("tab_6_report")
        self.pushButton_tab6_make_report = QtWidgets.QPushButton(self.tab_6_report)
        self.pushButton_tab6_make_report.setGeometry(QtCore.QRect(290, 190, 321, 211))
        self.pushButton_tab6_make_report.setObjectName("pushButton_tab6_make_report")
        self.tabWidget.addTab(self.tab_6_report, "")
        self.label_15_dont_touch = QtWidgets.QLabel(self.centralwidget)
        self.label_15_dont_touch.setGeometry(QtCore.QRect(200, 750, 261, 16))
        self.label_15_dont_touch.setObjectName("label_15_dont_touch")
        self.frame_2 = QtWidgets.QFrame(self.centralwidget)
        self.frame_2.setGeometry(QtCore.QRect(0, 0, 1161, 121))
        self.frame_2.setStyleSheet("background-color: rgb(45, 54, 76);\n"
                                   "")
        self.frame_2.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_2.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_2.setObjectName("frame_2")
        self.label_2_dont_touch = QtWidgets.QLabel(self.frame_2)
        self.label_2_dont_touch.setGeometry(QtCore.QRect(40, 10, 81, 81))
        font = QtGui.QFont()
        font.setPointSize(55)
        self.label_2_dont_touch.setFont(font)
        self.label_2_dont_touch.setObjectName("label_2_dont_touch")
        self.label_dont_touch = QtWidgets.QLabel(self.frame_2)
        self.label_dont_touch.setGeometry(QtCore.QRect(120, 30, 71, 51))
        font = QtGui.QFont()
        font.setFamily("Gabriola")
        font.setPointSize(45)
        self.label_dont_touch.setFont(font)
        self.label_dont_touch.setStyleSheet("color: rgb(113, 215, 78);")
        self.label_dont_touch.setObjectName("label_dont_touch")
        self.label_6_dont_touch = QtWidgets.QLabel(self.frame_2)
        self.label_6_dont_touch.setGeometry(QtCore.QRect(40, 90, 171, 16))
        self.label_6_dont_touch.setObjectName("label_6_dont_touch")
        self.label_you_are_welcom_name = QtWidgets.QLabel(self.frame_2)
        self.label_you_are_welcom_name.setGeometry(QtCore.QRect(1020, 40, 300, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_you_are_welcom_name.setFont(font)
        self.label_you_are_welcom_name.setObjectName("label_you_are_welcom_name")
        self.pushButton_de_login = QtWidgets.QPushButton(self.frame_2)
        self.pushButton_de_login.setGeometry(QtCore.QRect(1000, 70, 91, 21))
        self.pushButton_de_login.setAcceptDrops(False)
        self.pushButton_de_login.setStyleSheet("QPushButton:hover{\n"
                                               "background-color: rgb(113, 215, 78);\n"
                                               "border-radius: 13;\n"
                                               "}\n"
                                               "")
        self.pushButton_de_login.setObjectName("pushButton_de_login")
        self.dial = QtWidgets.QDial(self.frame_2)
        self.dial.setGeometry(QtCore.QRect(190, 0, 51, 71))
        self.dial.setObjectName("dial")
        self.frame = QtWidgets.QFrame(self.centralwidget)
        self.frame.setGeometry(QtCore.QRect(0, 120, 191, 681))
        self.frame.setStyleSheet("*{\n"
                                 "background-color: rgb(45, 54, 76);\n"
                                 "}\n"
                                 "QPushButton:hover{\n"
                                 "background-color: rgb(113, 215, 78);\n"
                                 "border-radius:13;\n"
                                 "}\n"
                                 "QPushButton{\n"
                                 "background-color: rgb(45, 54, 76);\n"
                                 "border-radius:3;\n"
                                 "border: 1px solid rgb(113,136,190);\n"
                                 "}\n"
                                 "\n"
                                 "")
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.pushButton_main_work_spase_go_part3 = QtWidgets.QPushButton(self.frame)
        self.pushButton_main_work_spase_go_part3.setGeometry(QtCore.QRect(0, 0, 191, 111))
        font = QtGui.QFont()
        font.setBold(True)
        font.setUnderline(False)
        font.setWeight(75)
        self.pushButton_main_work_spase_go_part3.setFont(font)
        self.pushButton_main_work_spase_go_part3.setObjectName("pushButton_main_work_spase_go_part3")
        self.pushButton_report_go_part4 = QtWidgets.QPushButton(self.frame)
        self.pushButton_report_go_part4.setGeometry(QtCore.QRect(0, 110, 191, 111))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_report_go_part4.setFont(font)
        self.pushButton_report_go_part4.setObjectName("pushButton_report_go_part4")
        self.pushButton_rating_workers_go_part5 = QtWidgets.QPushButton(self.frame)
        self.pushButton_rating_workers_go_part5.setGeometry(QtCore.QRect(0, 220, 191, 111))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_rating_workers_go_part5.setFont(font)
        self.pushButton_rating_workers_go_part5.setObjectName("pushButton_rating_workers_go_part5")
        self.pushButton_report_for_print_go_part6 = QtWidgets.QPushButton(self.frame)
        self.pushButton_report_for_print_go_part6.setGeometry(QtCore.QRect(0, 330, 191, 111))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        font.setStrikeOut(False)
        self.pushButton_report_for_print_go_part6.setFont(font)
        self.pushButton_report_for_print_go_part6.setObjectName("pushButton_report_for_print_go_part6")
        self.label = QtWidgets.QLabel(self.frame)
        self.label.setGeometry(QtCore.QRect(40, 540, 141, 16))
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.frame)
        self.label_2.setGeometry(QtCore.QRect(30, 520, 171, 20))
        self.label_2.setStyleSheet("")
        self.label_2.setObjectName("label_2")
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        # ADDD
        self.pushButton_main_work_spase_go_part3.setEnabled(False)
        self.pushButton_report_go_part4.setEnabled(False)
        self.pushButton_rating_workers_go_part5.setEnabled(False)
        self.pushButton_report_for_print_go_part6.setEnabled(False)
        self.pushButton_de_login.setVisible(False)

        self.pushButton_part2_registration_go_to_part2.clicked.connect(self.set_current_index2)
        self.pushButton_part2_go_to_part1.clicked.connect(self.set_current_index1)
        self.pushButton_main_work_spase_go_part3.clicked.connect(self.set_current_index3)
        self.pushButton_report_go_part4.clicked.connect(self.set_current_index4)
        self.pushButton_rating_workers_go_part5.clicked.connect(self.set_current_index5)
        self.pushButton_report_for_print_go_part6.clicked.connect(self.set_current_index6)
        self.pushButton_part2_registration.clicked.connect(self.registation_new_user)
        self.pushButton_de_login.clicked.connect(self.delogin_in_programm)
        self.pushButton_part2_autorization_join.clicked.connect(self.join_in_programm)

        # оснавная страница
        self.pushButton_part3_add_in_work_spase.clicked.connect(self.add_in_work_spase)
        self.pushButton_part3_delete.clicked.connect(self.delete_from_sql_movements)
        self.pushButton_part3_edit.clicked.connect(self.update_from_sql_movements)
        # страница 4
        self.pushButton_tab4_view_report.clicked.connect(self.viev_otchet_organization)

        # страница 5
        self.pushButton_tab5_view_report.clicked.connect(self.viev_otchet_product)

        ###  Тут заполнение комбобокса товары
        sql = 'SELECT name_product FROM product'
        result_sql = self.sql_return_table_in_list(sql)
        # Если нет в таблице продуктов добавим.
        if len(result_sql) == 0:
            sql = f'''INSERT INTO product (name_product) 
            VALUES ('Брелок'), ('Шнурок'), ('Шерсть'), ('Коробка'), ('Розетка'), ('Телевизор'), ('Приставка'), ('Кабель'), ('Пакеты'), ('Диван'), ('Пылесос');'''
            self.sql_just_execution_of_a_command(sql)
        sql = 'SELECT name_product FROM product'
        result_sql = self.sql_return_table_in_list(sql)
        for item in result_sql:
            self.comboBox_part3_product_name.addItem(str(item)[2:-3])
            self.comboBox_tab5_name_product.addItem(str(item)[2:-3])

        ###  Тут заполнение комбобокса организации
        sql = 'SELECT name_organization FROM organization'
        result_sql = self.sql_return_table_in_list(sql)
        if len(result_sql) == 0:
            sql = f'''INSERT INTO organization (name_organization) 
            VALUES ('ООО "ДулизДеталь"'), ('ООО "Монолес"'), ('ООО "Здоровые традиции"'), ('СООО "ЭксЛатСтрой"'), ('ООО "СанБизнесСтрой"'), ('ООО "Донадас плюс"'), ('ИП Падорин Сергей Евгеньевич'), ('ИП Калинин Александр Викторович'), ('ИП Лысенков Максим Валентинович'), ('ООО "Энерол"');'''
            self.sql_just_execution_of_a_command(sql)
        sql = 'SELECT name_organization FROM organization'
        result_sql = self.sql_return_table_in_list(sql)
        for item in result_sql:
            self.comboBox_tab3_name_orhanization.addItem(str(item)[2:-3])
            self.comboBox_tab4_name_organization.addItem(str(item)[2:-3])

        # установили название колонок
        self.tableWidget_part3_work_spase.setHorizontalHeaderLabels(
            ["№ перемещения", "Имя товара", "кол-во", "Название организации", "дата перемещения", "сотрудник"])
        self.tableWidget_tab4.setHorizontalHeaderLabels(
            ["№ перемещения", "Имя товара", "кол-во", "Название организации", "дата перемещения", "сотрудник"])
        self.tableWidget_tab5.setHorizontalHeaderLabels(
            ["имя товара", "остаток на складе", "количество перемещений", "крайнее перемещение"])
        self.tableWidget_tab5.resizeColumnsToContents()
        self.pushButton_tab6_make_report.clicked.connect(self.report_from_exel)

        # sdgjhsdgjhgsdjhsdggsdsghjgshjsdggsdhjsgjh
        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Store"))
        self.label_3_dont_touch.setText(_translate("MainWindow", "Добро пожаловать!"))
        self.label_4_dont_touch.setText(_translate("MainWindow", "Для работы войдите в систему."))
        self.lineEdit_part1_login.setPlaceholderText(_translate("MainWindow", "Логин"))
        self.lineEdit_part2_password.setPlaceholderText(_translate("MainWindow", "Пароль"))
        self.pushButton_part2_autorization_join.setText(_translate("MainWindow", "Войти"))
        self.pushButton_part2_registration_go_to_part2.setText(_translate("MainWindow", "Регистрация!"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("MainWindow", "Tab 1"))
        self.lineEdit_part2_input_password.setPlaceholderText(_translate("MainWindow", "Введите пароль"))
        self.lineEdit_part2_input_password_double.setPlaceholderText(_translate("MainWindow", "Повторите пароль"))
        self.lineEdit_part2_input_nameAndSurName.setPlaceholderText(_translate("MainWindow", "Псевдоним"))
        self.pushButton_part2_registration.setText(_translate("MainWindow", "Зарегистрироваться"))
        self.pushButton_part2_go_to_part1.setText(_translate("MainWindow", "назад!"))
        self.lineEdit_part2_input_login.setPlaceholderText(_translate("MainWindow", "Введите логин"))
        self.label_17_dont_touch.setText(_translate("MainWindow", "Регистрация"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("MainWindow", "Page"))
        self.pushButton_part3_add_in_work_spase.setText(_translate("MainWindow", "Добавить"))
        self.radioButton_part3_plus.setText(_translate("MainWindow", "Приход"))
        self.radioButton_part3_minus.setText(_translate("MainWindow", "Расход"))
        self.label_18_dont_touch.setText(_translate("MainWindow", "Товар"))
        self.label_19_dont_touch.setText(_translate("MainWindow", "Кол-во"))
        self.pushButton_part3_delete.setText(_translate("MainWindow", "Удалить"))
        self.pushButton_part3_edit.setText(_translate("MainWindow", "Редактировать"))
        self.comboBox_tab3_name_orhanization.setItemText(0, _translate("MainWindow", "Имя организации"))
        self.label_18_dont_touch_.setText(_translate("MainWindow", "Организация"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_3_), _translate("MainWindow", "Page"))
        self.pushButton_tab4_view_report.setText(_translate("MainWindow", "Посмотреть отчет"))
        self.comboBox_tab4_name_organization.setItemText(0, _translate("MainWindow", "Имя организации"))
        self.label_dont_touch_2.setText(
            _translate("MainWindow", "Выберите организацию отчет по которой хотите увидеть."))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_4_), _translate("MainWindow", "Page"))
        self.comboBox_tab5_name_product.setItemText(0, _translate("MainWindow", "Все товары"))
        self.pushButton_tab5_view_report.setText(_translate("MainWindow", "Показать отчет"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_5_), _translate("MainWindow", "Page"))
        self.pushButton_tab6_make_report.setText(_translate("MainWindow", "Подготовить отчет на печать!"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_6_report), _translate("MainWindow", "Page"))
        self.label_15_dont_touch.setText(_translate("MainWindow", "* соблюдайте порядок на рабочих местах!"))
        self.label_2_dont_touch.setText(_translate("MainWindow", "St"))
        self.label_dont_touch.setText(_translate("MainWindow", "ore"))
        self.label_6_dont_touch.setText(_translate("MainWindow", "Ваш товар - Наша забота!"))
        self.label_you_are_welcom_name.setText(_translate("MainWindow", "Добро пожаловать!"))
        self.pushButton_de_login.setText(_translate("MainWindow", "Выйти"))
        self.pushButton_main_work_spase_go_part3.setText(_translate("MainWindow", "Рабочая область"))
        self.pushButton_report_go_part4.setText(_translate("MainWindow", "Отчет"))
        self.pushButton_rating_workers_go_part5.setText(_translate("MainWindow", "Отчет по заданному товару"))
        self.pushButton_report_for_print_go_part6.setText(_translate("MainWindow", "Отчет на печать"))
        self.label.setText(_translate("MainWindow", "ПОЗ-41 Мохорев Э.О"))  ##  Изменил, опечатка была
        self.label_2.setText(_translate("MainWindow", "Выполнил студент группы"))


if __name__ == "__main__":
    import sys
    create_table_and_triggers() # cоздание таблиц и бд если её нет.
    app = QtWidgets.QApplication(sys.argv)
    app.setWindowIcon(QtGui.QIcon('store.ico'))
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
